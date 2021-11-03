from openpyxl import load_workbook
wb = load_workbook('Arginine_ExtractedData.xlsx')
lookup = wb['ReferenceSheet']
data = wb["cosmic_complete_mutation_argini"]
finalvalue="";

for row in data.iter_rows(): 
    for newrow in lookup.iter_rows():
            if row[10].value == newrow[0].value and row[14].value == newrow[1].value:
                if row[11].value==None:
                    row[11].value=newrow[2].value;
                    row[12].value=newrow[3].value;
                elif row[11].value!=None and row[13].value==None:
                    row[13].value=newrow[2].value;
                    row[14].value=newrow[3].value;
                elif row[13].value!=None and row[15].value==None:
                    row[15].value=newrow[2].value;
                    row[16].value=newrow[3].value;
                    print("idk")
                else:
                    row[17].value=newrow[2].value;
                    row[18].value=newrow[3].value;
    wb.save('Arginine_CodonsInserted.xlsx')

